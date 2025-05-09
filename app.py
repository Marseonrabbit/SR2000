from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, jsonify
import os
import pandas as pd
import requests
import re
import json
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from threading import Thread, Event
import uuid
import logging

# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.environ.get('SESSION_SECRET', os.environ.get('SECRET_KEY', 'your_secret_key'))
UPLOAD_FOLDER = os.environ.get('UPLOAD_FOLDER', 'uploads')
DOWNLOAD_FOLDER = os.environ.get('DOWNLOAD_FOLDER', 'downloads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['DOWNLOAD_FOLDER'] = DOWNLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['DOWNLOAD_FOLDER'], exist_ok=True)

# List of API keys to rotate through
API_KEYS = [
    "9abba08d9d6ef52d9919348996426443426b420a161779da1308816d8334eb36",
    "3dbc013b459f54f7936863d83bcc0cfc35b181159240fb2ac58d518361e249bc",
]
current_api_key_index = 0

jobs = {}

# Regex for private and broadcast IPs
PRIVATE_IP_REGEX = r'^(10\.|192\.168\.|172\.(1[6-9]|2[0-9]|3[0-1])\.|127\.0\.0\.1$)'
BROADCAST_IP_REGEX = r'^(\d{1,3}\.){3}255$'

def is_private_or_broadcast(ip):
    return re.match(PRIVATE_IP_REGEX, ip) or re.match(BROADCAST_IP_REGEX, ip)

def get_next_api_key():
    global current_api_key_index
    api_key = API_KEYS[current_api_key_index]
    current_api_key_index = (current_api_key_index + 1) % len(API_KEYS)
    return api_key

def get_country_name(country_code):
    if not country_code:
        return "Unknown Country"
    try:
        response = requests.get(f"https://restcountries.com/v3.1/alpha/{country_code}")
        if response.status_code == 200:
            country_data = response.json()
            return country_data[0]["name"]["common"]
        else:
            return "Unknown Country"
    except:
        return "Unknown Country"

def classify_reputation(score):
    if score == 0:
        return "Safe"
    elif 1 <= score <= 5:
        return "Neutral"
    elif score > 5:
        return "Poor"
    return "Unknown"

def get_ip_info(api_key, ip):
    url = f"https://www.virustotal.com/api/v3/ip_addresses/{ip}"
    headers = {"x-apikey": api_key}
    response = requests.get(url, headers=headers)
    if response.status_code == 401:
        return None, None, None, None
    if response.status_code == 200:
        data = response.json()
        attributes = data.get("data", {}).get("attributes", {})
        isp = attributes.get("as_owner", "Unknown ISP")
        country_code = attributes.get("country", "")
        country = get_country_name(country_code)
        last_analysis_stats = attributes.get("last_analysis_stats", {})
        malicious_count = last_analysis_stats.get("malicious", 0)
        total_engines = sum(last_analysis_stats.values()) if last_analysis_stats else 0
        reputation = classify_reputation(malicious_count)
        if total_engines == 0:
            malicious_score_str = "0/0"
        else:
            malicious_score_str = f"{malicious_count}/{total_engines}"
        return isp, country, reputation, malicious_score_str
    else:
        return "Error", "Error", "Error", "Error"

def get_comments(api_key, resource_type, resource_id, limit=5):
    url = f"https://www.virustotal.com/api/v3/{resource_type}/{resource_id}/comments"
    headers = {"x-apikey": api_key}
    params = {"limit": limit}
    response = requests.get(url, headers=headers, params=params)
    if response.status_code == 200:
        data = response.json()
        comments = data.get('data', [])
        top_comments = []
        for comment in comments:
            attributes = comment.get('attributes', {})
            top_comments.append({
                'user': attributes.get('user', 'Anonymous'),
                'date': attributes.get('date', ''),
                'comment': attributes.get('text', '')
            })
        return top_comments
    else:
        return []

def classify_hash_reputation(malicious_count, total_engines):
    if total_engines == 0:
        return "Unknown"
    malicious_percentage = (malicious_count / total_engines) * 100
    if malicious_percentage == 0:
        return "Safe"
    elif malicious_percentage <= 10:
        return "Low Risk"
    elif malicious_percentage <= 20:
        return "Moderate Risk"
    else:
        return "High Risk"

def get_hash_reputation(api_key, file_hash):
    url = f"https://www.virustotal.com/api/v3/files/{file_hash}"
    headers = {"x-apikey": api_key}
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        data = response.json()
        attributes = data.get("data", {}).get("attributes", {})
        last_analysis_stats = attributes.get("last_analysis_stats", {})
        malicious_count = last_analysis_stats.get("malicious", 0)
        total_engines = sum(last_analysis_stats.values()) if last_analysis_stats else 0
        reputation = classify_hash_reputation(malicious_count, total_engines)
        file_type = attributes.get("type_description", "Unknown")
        if total_engines == 0:
            malicious_score_str = "0/0"
        else:
            malicious_score_str = f"{malicious_count}/{total_engines}"
        return reputation, malicious_count, total_engines, file_type, malicious_score_str
    elif response.status_code == 404:
        return "Not Found", 0, 0, "Unknown", None
    elif response.status_code == 401:
        return "Invalid API Key", 0, 0, "Unknown", None
    else:
        return "Error", 0, 0, "Unknown", None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/lookup_ip', methods=['POST'])
def lookup_ip():
    ip = request.form['ip']
    if is_private_or_broadcast(ip):
        flash('The IP provided is private or a broadcast IP. Please insert a public IP for analysis.', 'warning')
        return redirect(url_for('index'))
    
    api_key = get_next_api_key()
    isp, country, reputation, malicious_score = get_ip_info(api_key, ip)
    if isp is None:
        flash('Invalid API Key.', 'danger')
        return redirect(url_for('index'))
    
    comments = get_comments(api_key, 'ip_addresses', ip, limit=5)
    result_data = {
        'ip': ip,
        'isp': isp,
        'country': country,
        'reputation': reputation,
        'malicious_score': malicious_score,
        'comments': comments
    }
    return render_template('single_ip_result.html', result=result_data)

@app.route('/bulk_upload')
def bulk_upload():
    return render_template('bulk_upload.html')

@app.route('/process_bulk_upload', methods=['POST'])
def process_bulk_upload():
    file = request.files.get('file')
    if not file or file.filename == '':
        flash('No selected file.', 'warning')
        return redirect(url_for('bulk_upload'))
    
    filename = secure_filename(file.filename)
    file_ext = os.path.splitext(filename)[1].lower()
    if file_ext not in ['.csv', '.xls', '.xlsx']:
        flash('Unsupported file format. Please upload a CSV or Excel file.', 'danger')
        return redirect(url_for('bulk_upload'))
    
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(file_path)
    flash('File successfully uploaded. Starting analysis...', 'success')
    
    job_id = str(uuid.uuid4())
    jobs[job_id] = {
        'status': 'Processing',
        'progress': 0,
        'result_file': None,
        'message': 'File successfully uploaded. Starting analysis...',
        'cancel_event': Event(),
        'is_single_ip': False
    }
    thread = Thread(target=process_file_thread, args=(job_id, file_path))
    thread.start()
    return redirect(url_for('bulk_progress', job_id=job_id))

def process_file_thread(job_id, file_path):
    try:
        api_key = get_next_api_key()
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext == '.csv':
            ip_df = pd.read_csv(file_path)
        else:
            ip_df = pd.read_excel(file_path)
        
        acceptable_columns = ['ip', 'ip_address', 'ipaddress', '"source_ip"']
        columns_lower = [col.lower() for col in ip_df.columns]
        ip_column = None
        for col in acceptable_columns:
            if col.lower() in columns_lower:
                ip_column = ip_df.columns[columns_lower.index(col.lower())]
                break
        
        if not ip_column:
            jobs[job_id]['status'] = 'Error'
            jobs[job_id]['message'] = "The uploaded file must contain a column named 'IP', 'ip_address', or 'IPAddress'."
            return
        
        total_ips = len(ip_df)
        if total_ips == 0:
            jobs[job_id]['status'] = 'Error'
            jobs[job_id]['message'] = "The uploaded file contains no IP addresses."
            return
        
        output_data = []
        for idx, ip in enumerate(ip_df[ip_column]):
            if jobs[job_id]['cancel_event'].is_set():
                jobs[job_id]['status'] = 'Canceled'
                jobs[job_id]['message'] = 'IP analysis canceled by user.'
                return
            
            if is_private_or_broadcast(ip):
                output_data.append({
                    "S. No.": idx + 1,
                    "IP": ip,
                    "ISP": "N/A",
                    "Country": "N/A",
                    "Reputation": "Private/Broadcast IP",
                    "Malicious Score": "N/A"
                })
                continue
            
            isp, country, reputation, malicious_score = get_ip_info(api_key, ip)
            if isp is None:
                jobs[job_id]['status'] = 'Error'
                jobs[job_id]['message'] = 'Invalid API Key.'
                return
            
            output_data.append({
                "S. No.": idx + 1,
                "IP": ip,
                "ISP": isp,
                "Country": country,
                "Reputation": reputation,
                "Malicious Score": malicious_score
            })
            jobs[job_id]['progress'] = int((idx + 1) / total_ips * 100)
        
        output_df = pd.DataFrame(output_data)
        output_file = os.path.join(app.config['DOWNLOAD_FOLDER'], f'{job_id}_result.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.title = "Source IP Details"
        
        for r_idx, row in enumerate(dataframe_to_rows(output_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                     top=Side(style="thin"), bottom=Side(style="thin"))
                cell.border = thin_border
                if r_idx == 1:
                    cell.fill = PatternFill(start_color="16365C", end_color="16365C", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
        
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        wb.save(output_file)
        jobs[job_id]['status'] = 'Completed'
        jobs[job_id]['result_file'] = output_file
        jobs[job_id]['message'] = 'IP analysis completed successfully.'
    except Exception as e:
        logger.error(f"Error processing file: {str(e)}")
        jobs[job_id]['status'] = 'Error'
        jobs[job_id]['message'] = f'Error processing file: {str(e)}'
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)

@app.route('/bulk_progress/<job_id>')
def bulk_progress(job_id):
    if job_id not in jobs:
        flash('Invalid job ID.', 'danger')
        return redirect(url_for('bulk_upload'))
    return render_template('bulk_lookup_progress.html', job_id=job_id)

@app.route('/progress/<job_id>')
def get_progress(job_id):
    if job_id in jobs:
        return jsonify({
            'status': jobs[job_id]['status'],
            'progress': jobs[job_id]['progress'],
            'message': jobs[job_id]['message']
        })
    else:
        return jsonify({'status': 'Error', 'message': 'Invalid job ID.'})

@app.route('/download/<job_id>')
def download_result(job_id):
    if job_id in jobs and jobs[job_id]['status'] == 'Completed':
        return send_file(jobs[job_id]['result_file'], as_attachment=True)
    else:
        flash('Result file not available.', 'warning')
        return redirect(url_for('bulk_upload'))

@app.route('/cancel_job/<job_id>', methods=['POST'])
def cancel_job(job_id):
    if job_id in jobs:
        jobs[job_id]['cancel_event'].set()
        flash('Job cancellation requested. Please wait...', 'warning')
    else:
        flash('Invalid job ID.', 'danger')
    
    if jobs.get(job_id, {}).get('is_single_ip', False):
        return redirect(url_for('hash_lookup'))
    elif 'hash' in job_id:
        return redirect(url_for('bulk_hash_upload'))
    else:
        return redirect(url_for('bulk_upload'))

@app.route('/hash_lookup')
def hash_lookup():
    return render_template('hash_lookup.html')

@app.route('/lookup_hash', methods=['POST'])
def lookup_hash():
    file_hash = request.form['file_hash'].strip()
    
    # Validate hash format
    hash_patterns = {
        'md5': r'^[a-fA-F0-9]{32}$',
        'sha1': r'^[a-fA-F0-9]{40}$',
        'sha256': r'^[a-fA-F0-9]{64}$'
    }
    
    valid_hash = False
    for hash_type, pattern in hash_patterns.items():
        if re.match(pattern, file_hash):
            valid_hash = True
            break
    
    if not valid_hash:
        flash('Invalid hash format. Please provide a valid MD5, SHA-1, or SHA-256 hash.', 'warning')
        return redirect(url_for('hash_lookup'))
    
    api_key = get_next_api_key()
    reputation, malicious_count, total_engines, file_type, malicious_score = get_hash_reputation(api_key, file_hash)
    
    if reputation == "Invalid API Key":
        flash('Invalid API Key.', 'danger')
        return redirect(url_for('hash_lookup'))
    
    comments = get_comments(api_key, 'files', file_hash, limit=5)
    
    result_data = {
        'file_hash': file_hash,
        'reputation': reputation,
        'malicious_count': malicious_count,
        'total_engines': total_engines,
        'file_type': file_type,
        'malicious_score': malicious_score,
        'comments': comments
    }
    
    return render_template('hash_result.html', result=result_data)

@app.route('/bulk_hash_upload')
def bulk_hash_upload():
    return render_template('bulk_hash_upload.html')

@app.route('/process_bulk_hash_upload', methods=['POST'])
def process_bulk_hash_upload():
    file = request.files.get('file')
    if not file or file.filename == '':
        flash('No selected file.', 'warning')
        return redirect(url_for('bulk_hash_upload'))
    
    filename = secure_filename(file.filename)
    file_ext = os.path.splitext(filename)[1].lower()
    if file_ext not in ['.csv', '.xls', '.xlsx']:
        flash('Unsupported file format. Please upload a CSV or Excel file.', 'danger')
        return redirect(url_for('bulk_hash_upload'))
    
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(file_path)
    flash('File successfully uploaded. Starting hash analysis...', 'success')
    
    job_id = f"hash_{str(uuid.uuid4())}"
    jobs[job_id] = {
        'status': 'Processing',
        'progress': 0,
        'result_file': None,
        'message': 'File successfully uploaded. Starting hash analysis...',
        'cancel_event': Event(),
        'is_hash': True
    }
    thread = Thread(target=process_hash_file_thread, args=(job_id, file_path))
    thread.start()
    return redirect(url_for('bulk_hash_progress', job_id=job_id))

def process_hash_file_thread(job_id, file_path):
    try:
        api_key = get_next_api_key()
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext == '.csv':
            hash_df = pd.read_csv(file_path)
        else:
            hash_df = pd.read_excel(file_path)
        
        # If there's only one column without a header, use it as-is
        if len(hash_df.columns) == 1 and hash_df.columns[0].startswith('Unnamed'):
            hash_column = hash_df.columns[0]
        else:
            # Try to find a column that contains hashes
            hash_patterns = {
                'md5': r'^[a-fA-F0-9]{32}$',
                'sha1': r'^[a-fA-F0-9]{40}$',
                'sha256': r'^[a-fA-F0-9]{64}$'
            }
            
            hash_column = None
            for col in hash_df.columns:
                # Check if most values in the column match hash patterns
                sample = hash_df[col].astype(str).iloc[:min(5, len(hash_df))].str.strip()
                matches = 0
                for value in sample:
                    for pattern in hash_patterns.values():
                        if re.match(pattern, value):
                            matches += 1
                            break
                
                if matches / len(sample) >= 0.6:  # If 60% or more match a hash pattern
                    hash_column = col
                    break
            
            # If no column is found, try common column names
            if not hash_column:
                acceptable_columns = ['hash', 'file_hash', 'filehash', 'md5', 'sha1', 'sha256']
                columns_lower = [col.lower() for col in hash_df.columns]
                for col in acceptable_columns:
                    if col.lower() in columns_lower:
                        hash_column = hash_df.columns[columns_lower.index(col.lower())]
                        break
        
        if not hash_column:
            jobs[job_id]['status'] = 'Error'
            jobs[job_id]['message'] = "Could not identify a column with hash values in the uploaded file."
            return
        
        total_hashes = len(hash_df)
        if total_hashes == 0:
            jobs[job_id]['status'] = 'Error'
            jobs[job_id]['message'] = "The uploaded file contains no hash values."
            return
        
        output_data = []
        for idx, file_hash in enumerate(hash_df[hash_column]):
            if jobs[job_id]['cancel_event'].is_set():
                jobs[job_id]['status'] = 'Canceled'
                jobs[job_id]['message'] = 'Hash analysis canceled by user.'
                return
            
            # Clean and validate hash format
            file_hash = str(file_hash).strip()
            valid_hash = False
            for pattern in hash_patterns.values():
                if re.match(pattern, file_hash):
                    valid_hash = True
                    break
            
            if not valid_hash:
                output_data.append({
                    "S. No.": idx + 1,
                    "Hash": file_hash,
                    "Hash Type": "Invalid Format",
                    "Reputation": "N/A",
                    "File Type": "N/A",
                    "Detection Ratio": "N/A"
                })
                continue
            
            hash_type = "MD5" if len(file_hash) == 32 else "SHA-1" if len(file_hash) == 40 else "SHA-256"
            reputation, malicious_count, total_engines, file_type, malicious_score = get_hash_reputation(api_key, file_hash)
            
            if reputation == "Invalid API Key":
                jobs[job_id]['status'] = 'Error'
                jobs[job_id]['message'] = 'Invalid API Key.'
                return
            
            output_data.append({
                "S. No.": idx + 1,
                "Hash": file_hash,
                "Hash Type": hash_type,
                "Reputation": reputation,
                "File Type": file_type,
                "Detection Ratio": malicious_score
            })
            jobs[job_id]['progress'] = int((idx + 1) / total_hashes * 100)
        
        output_df = pd.DataFrame(output_data)
        output_file = os.path.join(app.config['DOWNLOAD_FOLDER'], f'{job_id}_result.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.title = "Hash Analysis Results"
        
        for r_idx, row in enumerate(dataframe_to_rows(output_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                     top=Side(style="thin"), bottom=Side(style="thin"))
                cell.border = thin_border
                if r_idx == 1:
                    cell.fill = PatternFill(start_color="16365C", end_color="16365C", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
        
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        wb.save(output_file)
        jobs[job_id]['status'] = 'Completed'
        jobs[job_id]['result_file'] = output_file
        jobs[job_id]['message'] = 'Hash analysis completed successfully.'
    except Exception as e:
        logger.error(f"Error processing hash file: {str(e)}")
        jobs[job_id]['status'] = 'Error'
        jobs[job_id]['message'] = f'Error processing file: {str(e)}'
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)

@app.route('/bulk_hash_progress/<job_id>')
def bulk_hash_progress(job_id):
    if job_id not in jobs:
        flash('Invalid job ID.', 'danger')
        return redirect(url_for('bulk_hash_upload'))
    return render_template('bulk_hash_progress.html', job_id=job_id)

@app.route('/osint_framework')
def osint_framework():
    try:
        # Load OSINT framework data
        with open('static/data/osint_framework.json', 'r') as f:
            osint_data = json.load(f)
        return render_template('osint_framework.html', osint_data=osint_data)
    except FileNotFoundError:
        # If file not found, initialize with default structure
        osint_data = {
            "name": "OSINT Framework",
            "children": [
                {
                    "name": "Domain & IP Research",
                    "children": [
                        {"name": "VirusTotal", "url": "https://www.virustotal.com/"},
                        {"name": "AbuseIPDB", "url": "https://www.abuseipdb.com/"},
                        {"name": "Shodan", "url": "https://www.shodan.io/"},
                        {"name": "SecurityTrails", "url": "https://securitytrails.com/"}
                    ]
                },
                {
                    "name": "Email & Username Research",
                    "children": [
                        {"name": "Hunter.io", "url": "https://hunter.io/"},
                        {"name": "Email-Format", "url": "https://www.email-format.com/"},
                        {"name": "Have I Been Pwned", "url": "https://haveibeenpwned.com/"}
                    ]
                },
                {
                    "name": "Social Media Research",
                    "children": [
                        {"name": "Sherlock", "url": "https://github.com/sherlock-project/sherlock"},
                        {"name": "Social-Analyzer", "url": "https://github.com/qeeqbox/social-analyzer"}
                    ]
                }
            ]
        }
        
        # Create directory if it doesn't exist
        os.makedirs('static/data', exist_ok=True)
        
        # Save default data
        with open('static/data/osint_framework.json', 'w') as f:
            json.dump(osint_data, f, indent=2)
            
        return render_template('osint_framework.html', osint_data=osint_data)

@app.route('/google_dorks')
def google_dorks():
    try:
        # Load Google Dorks data
        with open('static/data/google_dorks.json', 'r') as f:
            dorks_data = json.load(f)
        
        # Get categories for filter
        categories = set()
        for dork in dorks_data:
            categories.add(dork.get('category', 'General'))
        
        categories = sorted(list(categories))
        
        return render_template('google_dorks.html', dorks=dorks_data, categories=categories)
    except FileNotFoundError:
        # If file not found, initialize with default data
        dorks_data = [
            {
                "category": "Sensitive Files",
                "query": "filetype:pdf site:example.com confidential",
                "description": "Finds confidential PDF files on a specific domain"
            },
            {
                "category": "Sensitive Files",
                "query": "filetype:xls OR filetype:xlsx intext:password",
                "description": "Finds Excel sheets containing passwords"
            },
            {
                "category": "Website Vulnerabilities",
                "query": "inurl:wp-content/uploads/",
                "description": "Finds WordPress uploads directory which might contain sensitive files"
            },
            {
                "category": "Website Vulnerabilities",
                "query": "intext:\"sql syntax near\" | intext:\"syntax error has occurred\" | intext:\"incorrect syntax near\" | intext:\"unexpected end of SQL command\" | intext:\"Warning: mysql_connect()\" | intext:\"Warning: mysql_query()\" | intext:\"Warning: pg_connect()\"",
                "description": "Finds pages with potential SQL errors"
            },
            {
                "category": "Exposed Credentials",
                "query": "intext:\"INDEX OF /\" intext:\".env\"",
                "description": "Finds exposed environment files which may contain secrets"
            },
            {
                "category": "Exposed Credentials",
                "query": "intitle:\"Index of\" intext:\".sql\"",
                "description": "Finds exposed SQL database dumps"
            },
            {
                "category": "Configuration Files",
                "query": "intitle:\"Index of\" intext:\"config.php\"",
                "description": "Finds exposed PHP configuration files"
            },
            {
                "category": "Configuration Files",
                "query": "intitle:\"Index of\" \"web.config\"",
                "description": "Finds exposed web.config files from Microsoft IIS servers"
            },
            {
                "category": "Network Devices",
                "query": "intitle:\"Login Page\" inurl:\"management\"",
                "description": "Finds login pages for network management interfaces"
            },
            {
                "category": "Network Devices",
                "query": "inurl:\"/sws/index.html\"",
                "description": "Finds Sonicwall routers login pages"
            },
            {
                "category": "Open Directories",
                "query": "intitle:\"Index of /\" \".git\"",
                "description": "Finds exposed Git repositories which may contain sensitive information"
            },
            {
                "category": "Exposed APIs",
                "query": "intext:\"api_key\" | intext:\"apiKey\" | intext:\"api key\" | intext:\"apikey\"",
                "description": "Finds exposed API keys in various formats"
            },
            {
                "category": "Exposed APIs",
                "query": "intitle:\"swagger ui\" inurl:api",
                "description": "Finds Swagger UI API documentation pages"
            },
            {
                "category": "IP Addresses",
                "query": "site:pastebin.com \"ip address\"",
                "description": "Finds IP addresses on Pastebin that might be part of data leaks"
            },
            {
                "category": "IP Addresses",
                "query": "\"ip camera\" | intitle:\"webcam\" inurl:\"/view/index.shtml\"",
                "description": "Finds exposed IP cameras"
            }
        ]
            
        # Create directory if it doesn't exist
        os.makedirs('static/data', exist_ok=True)
        
        # Save default data
        with open('static/data/google_dorks.json', 'w') as f:
            json.dump(dorks_data, f, indent=2)
        
        # Get categories for filter
        categories = set()
        for dork in dorks_data:
            categories.add(dork.get('category', 'General'))
        
        categories = sorted(list(categories))
            
        return render_template('google_dorks.html', dorks=dorks_data, categories=categories)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
